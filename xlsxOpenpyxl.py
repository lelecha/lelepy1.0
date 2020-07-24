#!/usr/bin/python
# -*- coding:utf8 -*-

import openpyxl


# 提取数据导入新的excel
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def helper(path):
    wb = openpyxl.Workbook()

    wb.save(path)

def cur_head(path):

    wb = openpyxl.load_workbook(path)
    sheet = wb.get_sheet_by_name('Sheet')


    wb.remove(sheet)
    wb.save(path)

def add_sheet(path, sheet_name):
    excel=load_workbook(path)
    excel.create_sheet(sheet_name)

    excel.save(path)
def add_content_to_sheet(path, sheet_name,content):
    index = len(content)
    excel = load_workbook(path)
    sheet = excel.get_sheet_by_name(sheet_name)

    for i in range(0, index):
        for j in range(0, len(content[i])):

            sheet.cell(row=i + 1, column=j + 1, value=str(content[i][j]))
    list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T','U','V','W']
    for i in list:
        sheet.column_dimensions[i].width = 20
    list2 = ['A1','B1','C1','D1','E1','F1','G1','H1','I1','J1','K1','L1','M1','N1','O1','P1','Q1','R1','S1','T1','U1','V1','W1','X1']
    for i in list2:
        sheet[i].font = openpyxl.styles.Font(bold = True)
        # sheet[i].fill = openpyxl.styles.PatternFill(fill_type=None, end_color='91E4CB')
        sheet[i].fill = openpyxl.styles.PatternFill("solid", fgColor="91E4CB")




    sheet.freeze_panes = 'A2'
    # for row in sheet.values:
    #     temp = list(row)
    #     num1 = int(sheet.cell(row + 1, 3).value)
    #     num2 = int(sheet.cell(row + 1, 4).value)
    #     sheet.cell(row + 1, 3).fill = PatternFill("solid", fgColor="1874CD")
    #
    #     sheet.cell(2, 4).fill = PatternFill("solid", fgColor="1874CD")


    for index, row in enumerate(sheet.rows):
        if index > 0:

            cell = row[4]
            if cell.value.isdigit():
                if int(cell.value) == 0:
                    cell.fill = PatternFill("solid", fgColor="FFFFCC")
                if int(cell.value) == 1:
                    cell.fill = PatternFill("solid", fgColor="91E4CB")
                if int(cell.value) == 2:
                    cell.fill = PatternFill("solid", fgColor="FFCCCC")
                if int(cell.value) == 3:
                    cell.fill = PatternFill("solid", fgColor="FFFF00")
                if int(cell.value) == 4:
                    cell.fill = PatternFill("solid", fgColor="99CCCC")
                if int(cell.value) == 5:
                    cell.fill = PatternFill("solid", fgColor="FFCC99")
                if int(cell.value) == 6:
                    cell.fill = PatternFill("solid", fgColor="CCCC99")
                if int(cell.value) == 7:
                    cell.fill = PatternFill("solid", fgColor="0099CC")
                if int(cell.value) == 8:
                    cell.fill = PatternFill("solid", fgColor="FFFFCC")
                if int(cell.value) == 9:
                    cell.fill = PatternFill("solid", fgColor="91E4CB")
                if int(cell.value) == 10:
                    cell.fill = PatternFill("solid", fgColor="FFCCCC")
                if int(cell.value) == 11:
                    cell.fill = PatternFill("solid", fgColor="FFFF00")
                if int(cell.value) == 12:
                    cell.fill = PatternFill("solid", fgColor="99CCCC")
                if int(cell.value) == 13:
                    cell.fill = PatternFill("solid", fgColor="FFCC99")
                if int(cell.value) == 14:
                    cell.fill = PatternFill("solid", fgColor="CCCC99")
                if int(cell.value) == 15:
                    cell.fill = PatternFill("solid", fgColor="0099CC")
                if int(cell.value) == 16:
                    cell.fill = PatternFill("solid", fgColor="FFFFCC")
                if int(cell.value) == 17:
                    cell.fill = PatternFill("solid", fgColor="91E4CB")
                if int(cell.value) == 18:
                    cell.fill = PatternFill("solid", fgColor="FFCCCC")
                if int(cell.value) == 19:
                    cell.fill = PatternFill("solid", fgColor="FFFF00")
                if int(cell.value) == 20:
                    cell.fill = PatternFill("solid", fgColor="99CCCC")
                if int(cell.value) == 21:
                    cell.fill = PatternFill("solid", fgColor="FFCC99")
                if int(cell.value) == 22:
                    cell.fill = PatternFill("solid", fgColor="CCCC99")
                if int(cell.value) == 23:
                    cell.fill = PatternFill("solid", fgColor="0099CC")
                if int(cell.value) == 24:
                    cell.fill = PatternFill("solid", fgColor="FFFFCC")
                if int(cell.value) == 25:
                    cell.fill = PatternFill("solid", fgColor="91E4CB")
                if int(cell.value) == 26:
                    cell.fill = PatternFill("solid", fgColor="FFCCCC")
                if int(cell.value) == 27:
                    cell.fill = PatternFill("solid", fgColor="FFFF00")
                if int(cell.value) == 28:
                    cell.fill = PatternFill("solid", fgColor="99CCCC")
                if int(cell.value) == 29:
                    cell.fill = PatternFill("solid", fgColor="FFCC99")
                if int(cell.value) == 30:
                    cell.fill = PatternFill("solid", fgColor="CCCC99")
                if int(cell.value) == 31:
                    cell.fill = PatternFill("solid", fgColor="0099CC")
                if int(cell.value) == 32:
                    cell.fill = PatternFill("solid", fgColor="FFCC99")
                if int(cell.value) == 33:
                    cell.fill = PatternFill("solid", fgColor="CCCC99")
                if int(cell.value) == 34:
                    cell.fill = PatternFill("solid", fgColor="0099CC")
                if int(cell.value) == 35:
                    cell.fill = PatternFill("solid", fgColor="FFFFCC")
                if int(cell.value) == 36:
                    cell.fill = PatternFill("solid", fgColor="91E4CB")
                if int(cell.value) == 37:
                    cell.fill = PatternFill("solid", fgColor="FFCCCC")
                if int(cell.value) == 38:
                    cell.fill = PatternFill("solid", fgColor="FFFF00")
                if int(cell.value) == 39:
                    cell.fill = PatternFill("solid", fgColor="99CCCC")
                if int(cell.value) == 40:
                    cell.fill = PatternFill("solid", fgColor="FFCC99")
                if int(cell.value) == 41:
                    cell.fill = PatternFill("solid", fgColor="CCCC99")
                if int(cell.value) == 42:
                    cell.fill = PatternFill("solid", fgColor="0099CC")

            cell = row[5]
            if cell.value.isdigit():
                if int(cell.value) == 0:
                    cell.fill = PatternFill("solid", fgColor="FFFFCC")
                if int(cell.value) == 1:
                    cell.fill = PatternFill("solid", fgColor="91E4CB")
                if int(cell.value) == 2:
                    cell.fill = PatternFill("solid", fgColor="FFCCCC")
                if int(cell.value) == 3:
                    cell.fill = PatternFill("solid", fgColor="FFFF00")
                if int(cell.value) == 4:
                    cell.fill = PatternFill("solid", fgColor="99CCCC")
                if int(cell.value) == 5:
                    cell.fill = PatternFill("solid", fgColor="FFCC99")
                if int(cell.value) == 6:
                    cell.fill = PatternFill("solid", fgColor="CCCC99")
                if int(cell.value) == 7:
                    cell.fill = PatternFill("solid", fgColor="0099CC")
                if int(cell.value) == 8:
                    cell.fill = PatternFill("solid", fgColor="FFFFCC")
                if int(cell.value) == 9:
                    cell.fill = PatternFill("solid", fgColor="91E4CB")
                if int(cell.value) == 10:
                    cell.fill = PatternFill("solid", fgColor="FFCCCC")
                if int(cell.value) == 11:
                    cell.fill = PatternFill("solid", fgColor="FFFF00")
                if int(cell.value) == 12:
                    cell.fill = PatternFill("solid", fgColor="99CCCC")
                if int(cell.value) == 13:
                    cell.fill = PatternFill("solid", fgColor="FFCC99")
                if int(cell.value) == 14:
                    cell.fill = PatternFill("solid", fgColor="CCCC99")
                if int(cell.value) == 15:
                    cell.fill = PatternFill("solid", fgColor="0099CC")
                if int(cell.value) == 16:
                    cell.fill = PatternFill("solid", fgColor="FFFFCC")
                if int(cell.value) == 17:
                    cell.fill = PatternFill("solid", fgColor="91E4CB")
                if int(cell.value) == 18:
                    cell.fill = PatternFill("solid", fgColor="FFCCCC")
                if int(cell.value) == 19:
                    cell.fill = PatternFill("solid", fgColor="FFFF00")
                if int(cell.value) == 20:
                    cell.fill = PatternFill("solid", fgColor="99CCCC")
                if int(cell.value) == 21:
                    cell.fill = PatternFill("solid", fgColor="FFCC99")
                if int(cell.value) == 22:
                    cell.fill = PatternFill("solid", fgColor="CCCC99")
                if int(cell.value) == 23:
                    cell.fill = PatternFill("solid", fgColor="0099CC")
                if int(cell.value) == 24:
                    cell.fill = PatternFill("solid", fgColor="FFFFCC")
                if int(cell.value) == 25:
                    cell.fill = PatternFill("solid", fgColor="91E4CB")
                if int(cell.value) == 26:
                    cell.fill = PatternFill("solid", fgColor="FFCCCC")
                if int(cell.value) == 27:
                    cell.fill = PatternFill("solid", fgColor="FFFF00")
                if int(cell.value) == 28:
                    cell.fill = PatternFill("solid", fgColor="99CCCC")
                if int(cell.value) == 29:
                    cell.fill = PatternFill("solid", fgColor="FFCC99")
                if int(cell.value) == 30:
                    cell.fill = PatternFill("solid", fgColor="CCCC99")
                if int(cell.value) == 31:
                    cell.fill = PatternFill("solid", fgColor="0099CC")
                if int(cell.value) == 32:
                    cell.fill = PatternFill("solid", fgColor="FFCC99")
                if int(cell.value) == 33:
                    cell.fill = PatternFill("solid", fgColor="CCCC99")
                if int(cell.value) == 34:
                    cell.fill = PatternFill("solid", fgColor="0099CC")
                if int(cell.value) == 35:
                    cell.fill = PatternFill("solid", fgColor="FFFFCC")
                if int(cell.value) == 36:
                    cell.fill = PatternFill("solid", fgColor="91E4CB")
                if int(cell.value) == 37:
                    cell.fill = PatternFill("solid", fgColor="FFCCCC")
                if int(cell.value) == 38:
                    cell.fill = PatternFill("solid", fgColor="FFFF00")
                if int(cell.value) == 39:
                    cell.fill = PatternFill("solid", fgColor="99CCCC")
                if int(cell.value) == 40:
                    cell.fill = PatternFill("solid", fgColor="FFCC99")
                if int(cell.value) == 41:
                    cell.fill = PatternFill("solid", fgColor="CCCC99")
                if int(cell.value) == 42:
                    cell.fill = PatternFill("solid", fgColor="0099CC")
    excel.save(path)




def write_excel_xlsx(path, sheet_name, value):
    # sheet = load_workbook('path')
    # sheet.create_sheet(sheet_name)

    index = len(value)
    workbook = openpyxl.Workbook()

    # sheet = workbook.create_sheet(sheet_name)
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.freeze_panes = 'A2'
    for i in range(0, index):
        for j in range(0, len(value[i])):

            sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
    list = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T']
    for i in list:
        sheet.column_dimensions[i].width = 20

    list2 = ['A1','B1','C1','D1','E1','F1','G1','H1','I1','J1','K1','L1','M1','N1','O1','P1','Q1','R1','S1','T1','U1','V1','W1','X1','Y1']
    for i in list2:
        sheet[i].font = openpyxl.styles.Font(bold = True)
        # sheet[i].fill = openpyxl.styles.PatternFill(fill_type=None, end_color='91E4CB')
        sheet[i].fill = openpyxl.styles.PatternFill("solid", fgColor="91E4CB")
    # for row in sheet.values:
    #     temp = list(row)
    #     num1 = int(sheet.cell(row + 1, 3).value)
    #     num2 = int(sheet.cell(row + 1, 4).value)
    #     sheet.cell(row + 1, 3).fill = PatternFill("solid", fgColor="1874CD")
    #
    #     sheet.cell(2, 4).fill = PatternFill("solid", fgColor="1874CD")

    for index, row in enumerate(sheet.rows):
        if index > 0:
            cell = row[3]
            if cell.value.isdigit():
                if int(cell.value) == 0:
                    cell.fill = PatternFill("solid", fgColor="FFFFCC")
                if int(cell.value) == 1:
                    cell.fill = PatternFill("solid", fgColor="91E4CB")
                if int(cell.value) == 2:
                    cell.fill = PatternFill("solid", fgColor="FFCCCC")
                if int(cell.value) == 3:
                    cell.fill = PatternFill("solid", fgColor="FFFF00")
                if int(cell.value) == 4:
                    cell.fill = PatternFill("solid", fgColor="99CCCC")
                if int(cell.value) == 5:
                    cell.fill = PatternFill("solid", fgColor="FFCC99")
                if int(cell.value) == 6:
                    cell.fill = PatternFill("solid", fgColor="CCCC99")
                if int(cell.value) == 7:
                    cell.fill = PatternFill("solid", fgColor="0099CC")
                if int(cell.value) == 8:
                    cell.fill = PatternFill("solid", fgColor="FFFFCC")
                if int(cell.value) == 9:
                    cell.fill = PatternFill("solid", fgColor="91E4CB")
                if int(cell.value) == 10:
                    cell.fill = PatternFill("solid", fgColor="FFCCCC")
                if int(cell.value) == 11:
                    cell.fill = PatternFill("solid", fgColor="FFFF00")
                if int(cell.value) == 12:
                    cell.fill = PatternFill("solid", fgColor="99CCCC")
                if int(cell.value) == 13:
                    cell.fill = PatternFill("solid", fgColor="FFCC99")
                if int(cell.value) == 14:
                    cell.fill = PatternFill("solid", fgColor="CCCC99")
                if int(cell.value) == 15:
                    cell.fill = PatternFill("solid", fgColor="0099CC")
                if int(cell.value) == 16:
                    cell.fill = PatternFill("solid", fgColor="FFFFCC")
                if int(cell.value) == 17:
                    cell.fill = PatternFill("solid", fgColor="91E4CB")
                if int(cell.value) == 18:
                    cell.fill = PatternFill("solid", fgColor="FFCCCC")
                if int(cell.value) == 19:
                    cell.fill = PatternFill("solid", fgColor="FFFF00")
                if int(cell.value) == 20:
                    cell.fill = PatternFill("solid", fgColor="99CCCC")
                if int(cell.value) == 21:
                    cell.fill = PatternFill("solid", fgColor="FFCC99")
                if int(cell.value) == 22:
                    cell.fill = PatternFill("solid", fgColor="CCCC99")
                if int(cell.value) == 23:
                    cell.fill = PatternFill("solid", fgColor="0099CC")
                if int(cell.value) == 24:
                    cell.fill = PatternFill("solid", fgColor="FFFFCC")
                if int(cell.value) == 25:
                    cell.fill = PatternFill("solid", fgColor="91E4CB")
                if int(cell.value) == 26:
                    cell.fill = PatternFill("solid", fgColor="FFCCCC")
                if int(cell.value) == 27:
                    cell.fill = PatternFill("solid", fgColor="FFFF00")
                if int(cell.value) == 28:
                    cell.fill = PatternFill("solid", fgColor="99CCCC")
                if int(cell.value) == 29:
                    cell.fill = PatternFill("solid", fgColor="FFCC99")
                if int(cell.value) == 30:
                    cell.fill = PatternFill("solid", fgColor="CCCC99")
                if int(cell.value) == 31:
                    cell.fill = PatternFill("solid", fgColor="0099CC")

            cell = row[4]
            if cell.value.isdigit():
                if int(cell.value) == 0:
                    cell.fill = PatternFill("solid", fgColor="FFFFCC")
                if int(cell.value) == 1:
                    cell.fill = PatternFill("solid", fgColor="91E4CB")
                if int(cell.value) == 2:
                    cell.fill = PatternFill("solid", fgColor="FFCCCC")
                if int(cell.value) == 3:
                    cell.fill = PatternFill("solid", fgColor="FFFF00")
                if int(cell.value) == 4:
                    cell.fill = PatternFill("solid", fgColor="99CCCC")
                if int(cell.value) == 5:
                    cell.fill = PatternFill("solid", fgColor="FFCC99")
                if int(cell.value) == 6:
                    cell.fill = PatternFill("solid", fgColor="CCCC99")
                if int(cell.value) == 7:
                    cell.fill = PatternFill("solid", fgColor="0099CC")
                if int(cell.value) == 8:
                    cell.fill = PatternFill("solid", fgColor="FFFFCC")
                if int(cell.value) == 9:
                    cell.fill = PatternFill("solid", fgColor="91E4CB")
                if int(cell.value) == 10:
                    cell.fill = PatternFill("solid", fgColor="FFCCCC")
                if int(cell.value) == 11:
                    cell.fill = PatternFill("solid", fgColor="FFFF00")
                if int(cell.value) == 12:
                    cell.fill = PatternFill("solid", fgColor="99CCCC")
                if int(cell.value) == 13:
                    cell.fill = PatternFill("solid", fgColor="FFCC99")
                if int(cell.value) == 14:
                    cell.fill = PatternFill("solid", fgColor="CCCC99")
                if int(cell.value) == 15:
                    cell.fill = PatternFill("solid", fgColor="0099CC")
                if int(cell.value) == 16:
                    cell.fill = PatternFill("solid", fgColor="FFFFCC")
                if int(cell.value) == 17:
                    cell.fill = PatternFill("solid", fgColor="91E4CB")
                if int(cell.value) == 18:
                    cell.fill = PatternFill("solid", fgColor="FFCCCC")
                if int(cell.value) == 19:
                    cell.fill = PatternFill("solid", fgColor="FFFF00")
                if int(cell.value) == 20:
                    cell.fill = PatternFill("solid", fgColor="99CCCC")
                if int(cell.value) == 21:
                    cell.fill = PatternFill("solid", fgColor="FFCC99")
                if int(cell.value) == 22:
                    cell.fill = PatternFill("solid", fgColor="CCCC99")
                if int(cell.value) == 23:
                    cell.fill = PatternFill("solid", fgColor="0099CC")
                if int(cell.value) == 24:
                    cell.fill = PatternFill("solid", fgColor="FFFFCC")
                if int(cell.value) == 25:
                    cell.fill = PatternFill("solid", fgColor="91E4CB")
                if int(cell.value) == 26:
                    cell.fill = PatternFill("solid", fgColor="FFCCCC")
                if int(cell.value) == 27:
                    cell.fill = PatternFill("solid", fgColor="FFFF00")
                if int(cell.value) == 28:
                    cell.fill = PatternFill("solid", fgColor="99CCCC")
                if int(cell.value) == 29:
                    cell.fill = PatternFill("solid", fgColor="FFCC99")
                if int(cell.value) == 30:
                    cell.fill = PatternFill("solid", fgColor="CCCC99")
                if int(cell.value) == 31:
                    cell.fill = PatternFill("solid", fgColor="0099CC")













    # for i in sheet.columns:
    #     sheet.column_dimensions[i].width = 20.0
    workbook.save(path)
    print("success")


def write_excel_newLine(path, sheet_name, value):  # 这里的value是新的内容
    # index = len(value)
    list = read_excel_xlsx(path, sheet_name)
    for i in value:
        list.append(i)
    write_excel_xlsx(path, sheet_name, list)
    # for i in range(0, index):
    #     for j in range(0, len(value[i])):
    #         sheet.cell(row=i + 1, column=j + 1, value=str(value[i][j]))
    # workbook.save(path)
    print("insert new line into excel")


def read_excel_xlsx(path):
    workbook = openpyxl.load_workbook(path)
    sheets = workbook.get_sheet_names()
    sheets = list(sheets)
    lists = []
    for i in sheets:
        onesheet = []
        sheet = workbook[i]
        curlist = list(sheet.values)
        if curlist != None and len(curlist) != 0:
            del(curlist[0])
        for row in curlist:
            temp = list(row)
            for j in range(len(temp)):
                temp[j] = str(temp[j])
            onesheet.append(temp)
        # del (onesheet[0])
        lists.append(onesheet)
    lists.insert(0,sheets)



    return lists
    # for j in row:
    #     print(j)

    # list = tuple(sheet.rows)
    # print('111')

    # for row in sheet.rows:
    #     for cell in row:
    #         print(cell.value, "\t")
    #     print()

# book_name_xlsx = 'text2.xlsx'
#
# sheet_name_xlsx = 'xlsxformtest'
#
# value3 = [["姓名", "性别", "年龄", "城市", "职业"],
#           ["111", "女", "66", "石家庄", "运维工程师"],
#           ["222", "男", "55", "南京", "饭店老板"],
#           ["333", "女", "27", "苏州", "保安"], ]
#
# write_excel_xlsx(book_name_xlsx, sheet_name_xlsx, value3)
